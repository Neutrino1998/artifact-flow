#!/usr/bin/env python3
"""工作簿公式静态体检 —— 本环境不能重算公式,这是写完公式后唯一的自动检查。

用法:
    python check_formulas.py 文件.xlsx

检查项(输出 JSON,issues 非空 → 退出码 1):
  broken_ref    公式文本里已有 #REF!(引用在编辑中被破坏)
  unknown_sheet 公式引用了不存在的工作表
  cached_error  单元格缓存值是 Excel 错误(#DIV/0! 等 —— 仅对 Excel 算过的
                文件有意义;本环境新写的公式没有缓存值,查不到属正常)

查不出的:循环引用、类型错误、逻辑错误 —— 公式正确性最终以 Excel 打开重算为准。
"""

import json
import re
import sys

from openpyxl import load_workbook

# 带引号表名支持 '' 转义('John''s Data'!)
_SHEET_REF_RE = re.compile(r"(?:'((?:[^']|'')+)'|([A-Za-z0-9_一-鿿]+))!")
_STRING_LITERAL_RE = re.compile(r'"(?:[^"]|"")*"')
# 裸外部工作簿引用([1]Sheet1!)整 token 剥掉 —— 不能用 lookbehind 挡:
# 引擎在 ] 后一位重试会匹配出残名(heet1)
_EXTERNAL_REF_RE = re.compile(r"\[[^\]]+\][A-Za-z0-9_一-鿿]*!")
_ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}


def main():
    if len(sys.argv) != 2:
        raise SystemExit(__doc__)
    path = sys.argv[1]
    issues = []

    wb = load_workbook(path, data_only=False)
    sheet_names = set(wb.sheetnames)
    n_formulas = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue
                n_formulas += 1
                formula = str(cell.value)
                loc = f"{ws.title}!{cell.coordinate}"
                if "#REF!" in formula:
                    issues.append({"cell": loc, "kind": "broken_ref", "formula": formula[:80]})
                # 先剥字符串字面量("完成!" 里的 ! 不是表引用)与裸外部引用
                scannable = _STRING_LITERAL_RE.sub('""', formula)
                scannable = _EXTERNAL_REF_RE.sub("", scannable)
                for m in _SHEET_REF_RE.finditer(scannable):
                    ref = (m.group(1) or m.group(2)).replace("''", "'")
                    if "[" in ref:      # 外部工作簿引用('[Book1]Sheet1'!),不查
                        continue
                    # 纯行列引用(如 A1)会被误捕为表名,只查真名单外且非单元格样式的
                    if ref not in sheet_names and not re.fullmatch(r"[A-Za-z]{1,3}[0-9]*", ref):
                        issues.append({"cell": loc, "kind": "unknown_sheet",
                                       "formula": formula[:80], "sheet": ref})

    wb2 = load_workbook(path, data_only=True)
    for ws in wb2.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in _ERROR_VALUES:
                    issues.append({"cell": f"{ws.title}!{cell.coordinate}",
                                   "kind": "cached_error", "value": cell.value})

    print(json.dumps({"formulas": n_formulas, "issues": issues},
                     ensure_ascii=False, indent=1))
    sys.exit(1 if issues else 0)


if __name__ == "__main__":
    main()
