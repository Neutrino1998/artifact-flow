#!/usr/bin/env python3
"""In-sandbox ENOSYS probe (plan §B — the headline risk).

Run UNDER runsc, e.g. via run-all.sh, or directly:
    docker run --rm --runtime=runsc --network=none \\
        -v "$PWD/sandbox/verify:/opt/verify:ro" \\
        artifactflow-sandbox:latest python3 /opt/verify/verify-enosys.py

Each library is exercised with a REAL workload (not echo). On failure it prints
the lib + the failing op, and flags errno==ENOSYS(38) specially — that's the
gVisor Sentry "syscall not implemented" gap, i.e. the signal to consider the
Firecracker fallback for that lib. C-extension libs (numpy/pandas/matplotlib/
Pillow) carry the real risk; openpyxl/pypdf are pure-Python and labeled as such
so a green there does not mask a C-ext failure.

Probe order matters: matplotlib(PDF) writes the fixture that pypdf then parses —
fixtures are self-generated, nothing binary is carried onto the air-gapped node.
"""
import errno
import os
import sys
import tempfile
import traceback

WORK = tempfile.mkdtemp(prefix="enosys-")
results = []  # (name, kind, ok, detail)


def probe(name, kind, fn):
    try:
        fn()
        results.append((name, kind, True, ""))
    except Exception as e:  # noqa: BLE001 — we want to catch and classify everything
        en = getattr(e, "errno", None)
        if en == errno.ENOSYS:
            detail = f"ENOSYS(38) — gVisor Sentry syscall gap: {e!r}"
        else:
            last = traceback.format_exc().strip().splitlines()[-1]
            detail = f"{type(e).__name__}: {e} | {last}"
        results.append((name, kind, False, detail))


def t_numpy():
    import numpy as np

    a = np.random.rand(800, 800)
    float((a @ a.T).sum())


def t_pandas():
    import pandas as pd

    df = pd.DataFrame({"a": range(10000)})
    df["b"] = df["a"] ** 2
    df.groupby(df["a"] % 7)["b"].sum()


def t_matplotlib_png():
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots()
    ax.plot([1, 2, 3], [4, 5, 6])
    fig.savefig(os.path.join(WORK, "plot.png"))
    plt.close(fig)


def t_matplotlib_pdf():
    # Doubles as the fixture pypdf parses below (vector PDF — no LaTeX needed).
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots()
    ax.set_title("enosys-probe")
    ax.plot([0, 1], [0, 1])
    fig.savefig(os.path.join(WORK, "doc.pdf"))
    plt.close(fig)


def t_pillow():
    from PIL import Image

    p = os.path.join(WORK, "img.png")
    Image.new("RGB", (64, 48), (123, 222, 64)).save(p)
    Image.open(p).load()


def t_openpyxl():
    from openpyxl import Workbook, load_workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "x", 42
    p = os.path.join(WORK, "s.xlsx")
    wb.save(p)
    assert load_workbook(p).active["B1"].value == 42


def t_pypdf():
    from pypdf import PdfReader

    p = os.path.join(WORK, "doc.pdf")
    if not os.path.exists(p):
        raise RuntimeError("fixture doc.pdf missing (matplotlib_pdf must pass first)")
    _ = PdfReader(p).pages[0]


probe("numpy", "C-ext", t_numpy)
probe("pandas", "C-ext", t_pandas)
probe("matplotlib(PNG)", "C-ext", t_matplotlib_png)
probe("matplotlib(PDF)", "C-ext", t_matplotlib_pdf)
probe("Pillow", "C-ext", t_pillow)
probe("openpyxl", "pure-Python", t_openpyxl)
probe("pypdf", "pure-Python", t_pypdf)

print(f"\nENOSYS probe — python {sys.version.split()[0]}\n")
width = max(len(r[0]) for r in results)
nfail = 0
for name, kind, ok, detail in results:
    status = "PASS" if ok else "FAIL"
    line = f"  [{status}] {name:<{width}}  ({kind})"
    if not ok:
        line += f"\n         {detail}"
        nfail += 1
    print(line)

summary = f"\n{len(results) - nfail}/{len(results)} passed"
if nfail:
    summary += f" — {nfail} FAILED (any C-ext failure = Firecracker-fallback signal)"
print(summary)
sys.exit(1 if nfail else 0)
