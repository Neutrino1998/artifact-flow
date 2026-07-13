#!/usr/bin/env python3
"""Stable LibreOffice entry point for ArtifactFlow's offline sandbox.

The model-facing contract is intentionally small:

    artifactflow-office convert INPUT OUTPUT
    artifactflow-office render INPUT OUTDIR [--pages 1-5,8]
    artifactflow-office recalc INPUT.xlsx OUTPUT.xlsx

Every LibreOffice invocation gets an isolated writable user profile. Commands
fail loudly when LibreOffice exits unsuccessfully or does not create the
expected file; successful commands print one compact JSON object.
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import signal
import subprocess
import sys
import tempfile
from pathlib import Path

CLI_VERSION = "1.0"
OFFICE_TIMEOUT_SECONDS = 120
RENDER_DPI = 150
MAX_WORKBOOK_SCAN_CELLS = 500_000
MAX_REPORTED_FORMULA_ERRORS = 100
MAX_RENDER_PAGES = 100
MAX_RENDER_PIXELS_PER_PAGE = 40_000_000

_CONVERT_FILTERS = {
    ".docx": 'docx:Office Open XML Text',
    ".odt": "odt:writer8",
    ".pptx": 'pptx:Impress MS PowerPoint 2007 XML',
    ".odp": "odp:impress8",
    ".xlsx": 'xlsx:Calc MS Excel 2007 XML',
    ".ods": "ods:calc8",
    ".pdf": "pdf",
    ".html": "html",
    ".txt": "txt:Text",
    ".csv": "csv:Text - txt - csv (StarCalc)",
}


class OfficeError(RuntimeError):
    pass


def _emit(payload: dict, *, stream=sys.stdout) -> None:
    print(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), file=stream)


def _soffice() -> str:
    executable = shutil.which("soffice") or shutil.which("libreoffice")
    if not executable:
        raise OfficeError("LibreOffice executable not found in sandbox image")
    return executable


def _run_process(command: list[str], *, env: dict[str, str]) -> subprocess.CompletedProcess[str]:
    process = subprocess.Popen(
        command,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        env=env,
        start_new_session=True,
    )
    try:
        stdout, stderr = process.communicate(timeout=OFFICE_TIMEOUT_SECONDS)
    except subprocess.TimeoutExpired:
        try:
            os.killpg(process.pid, signal.SIGKILL)
        except ProcessLookupError:
            pass
        stdout, stderr = process.communicate()
        detail = (stderr or stdout).strip()[-2000:]
        raise OfficeError(
            f"LibreOffice timed out after {OFFICE_TIMEOUT_SECONDS}s"
            + (f": {detail}" if detail else "")
        )
    result = subprocess.CompletedProcess(command, process.returncode, stdout, stderr)
    if result.returncode != 0:
        detail = (result.stderr or result.stdout).strip()[-2000:]
        raise OfficeError(
            f"LibreOffice exited with code {result.returncode}"
            + (f": {detail}" if detail else "")
        )
    return result


def _office_env(home: Path) -> dict[str, str]:
    env = os.environ.copy()
    env.update(
        {
            "HOME": str(home),
            "XDG_CACHE_HOME": str(home / ".cache"),
            "XDG_CONFIG_HOME": str(home / ".config"),
            "XDG_DATA_HOME": str(home / ".local" / "share"),
            "SAL_USE_VCLPLUGIN": "svp",
        }
    )
    for key in ("HOME", "XDG_CACHE_HOME", "XDG_CONFIG_HOME", "XDG_DATA_HOME"):
        Path(env[key]).mkdir(parents=True, exist_ok=True)
    return env


def _conversion_spec(suffix: str) -> str:
    try:
        return _CONVERT_FILTERS[suffix.lower()]
    except KeyError as exc:
        supported = ", ".join(sorted(_CONVERT_FILTERS))
        raise OfficeError(f"unsupported output extension {suffix!r}; supported: {supported}") from exc


def _atomic_copy(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    staging = target.with_name(f".{target.name}.artifactflow-office-{os.getpid()}")
    try:
        shutil.copy2(source, staging)
        os.replace(staging, target)
    finally:
        staging.unlink(missing_ok=True)


def _convert_to_temp(source: Path, suffix: str, work: Path) -> tuple[Path, subprocess.CompletedProcess[str]]:
    outdir = work / "output"
    profile = work / "profile"
    home = work / "home"
    outdir.mkdir(parents=True)
    profile.mkdir(parents=True)
    command = [
        _soffice(),
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--nofirststartwizard",
        "--norestore",
        f"-env:UserInstallation={profile.resolve().as_uri()}",
        "--convert-to",
        _conversion_spec(suffix),
        "--outdir",
        str(outdir),
        str(source),
    ]
    result = _run_process(command, env=_office_env(home))
    candidates = sorted(p for p in outdir.iterdir() if p.is_file() and p.suffix.lower() == suffix)
    if len(candidates) != 1:
        listing = [p.name for p in sorted(outdir.iterdir())]
        detail = (result.stderr or result.stdout).strip()[-1000:]
        raise OfficeError(
            f"LibreOffice created {len(candidates)} {suffix} file(s), expected one; "
            f"output={listing}" + (f"; log={detail}" if detail else "")
        )
    return candidates[0], result


def convert(source: Path, target: Path) -> dict:
    source = source.resolve()
    target = target.resolve()
    if not source.is_file():
        raise OfficeError(f"input is not a file: {source}")
    if source == target:
        raise OfficeError("input and output must be different paths")
    suffix = target.suffix.lower()
    with tempfile.TemporaryDirectory(prefix="artifactflow-office-") as tmp:
        converted, result = _convert_to_temp(source, suffix, Path(tmp))
        _atomic_copy(converted, target)
    payload = {
        "ok": True,
        "operation": "convert",
        "input": str(source),
        "output": str(target),
        "bytes": target.stat().st_size,
    }
    warning = result.stderr.strip()
    if warning:
        payload["warning"] = warning[-1000:]
    return payload


def _parse_pages(expr: str | None, page_count: int) -> list[int]:
    if not expr:
        if page_count > MAX_RENDER_PAGES:
            raise OfficeError(
                f"document has {page_count} pages; select at most {MAX_RENDER_PAGES} with --pages"
            )
        return list(range(page_count))
    selected: set[int] = set()
    for raw_part in expr.split(","):
        part = raw_part.strip()
        if not part:
            continue
        try:
            if "-" in part:
                start, end = (int(value) for value in part.split("-", 1))
                if start > end:
                    raise ValueError
                selected.update(range(start, end + 1))
            else:
                selected.add(int(part))
        except ValueError as exc:
            raise OfficeError(f"invalid page selection: {part!r}") from exc
    if not selected or min(selected) < 1 or max(selected) > page_count:
        raise OfficeError(f"page selection must be within 1-{page_count}")
    if len(selected) > MAX_RENDER_PAGES:
        raise OfficeError(
            f"requested {len(selected)} pages; render at most {MAX_RENDER_PAGES} pages per call"
        )
    return [number - 1 for number in sorted(selected)]


def render(source: Path, outdir: Path, pages: str | None) -> dict:
    source = source.resolve()
    outdir = outdir.resolve()
    if not source.is_file():
        raise OfficeError(f"input is not a file: {source}")
    outdir.mkdir(parents=True, exist_ok=True)
    stale = sorted(p.name for p in outdir.glob("page-*.png"))
    if stale:
        raise OfficeError(f"output directory already contains rendered pages: {stale[:5]}")

    try:
        import pypdfium2 as pdfium
    except ImportError as exc:
        raise OfficeError("pypdfium2 is not installed in sandbox image") from exc

    with tempfile.TemporaryDirectory(prefix="artifactflow-office-render-") as tmp:
        work = Path(tmp)
        if source.suffix.lower() == ".pdf":
            pdf_path = source
            converted = False
        else:
            pdf_path, _ = _convert_to_temp(source, ".pdf", work)
            converted = True

        document = pdfium.PdfDocument(str(pdf_path))
        try:
            selected = _parse_pages(pages, len(document))
            rendered_files = []
            scale = RENDER_DPI / 72
            for page_index in selected:
                page = document[page_index]
                try:
                    width, height = page.get_size()
                    pixels = width * scale * height * scale
                    if pixels > MAX_RENDER_PIXELS_PER_PAGE:
                        raise OfficeError(
                            f"page {page_index + 1} would render {int(pixels):,} pixels; "
                            "page dimensions are too large"
                        )
                    image = page.render(scale=scale).to_pil()
                    try:
                        output = work / f"page-{page_index + 1}.png"
                        image.save(output, format="PNG")
                        rendered_files.append(output)
                    finally:
                        image.close()
                finally:
                    page.close()
        finally:
            document.close()

        outputs = []
        for rendered_file in rendered_files:
            output = outdir / rendered_file.name
            _atomic_copy(rendered_file, output)
            outputs.append(str(output))

    return {
        "ok": True,
        "operation": "render",
        "input": str(source),
        "converted_to_pdf": converted,
        "dpi": RENDER_DPI,
        "pages": len(outputs),
        "outputs": outputs,
    }


def _scan_workbook(path: Path, *, data_only: bool) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise OfficeError("openpyxl is not installed in sandbox image") from exc

    formulas = 0
    errors = []
    error_count = 0
    scanned = 0
    truncated = False
    workbook = load_workbook(path, data_only=data_only, read_only=True)
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    scanned += 1
                    if scanned > MAX_WORKBOOK_SCAN_CELLS:
                        truncated = True
                        break
                    if cell.value is None:
                        continue
                    if not data_only and cell.data_type == "f":
                        formulas += 1
                    if data_only and cell.data_type == "e":
                        error_count += 1
                        if len(errors) < MAX_REPORTED_FORMULA_ERRORS:
                            errors.append({
                                "cell": f"{worksheet.title}!{cell.coordinate}",
                                "value": cell.value,
                            })
                if truncated:
                    break
            if truncated:
                break
    finally:
        workbook.close()
    return {
        "formulas": formulas,
        "cached_errors": errors,
        "cached_error_count": error_count,
        "scanned_cells": min(scanned, MAX_WORKBOOK_SCAN_CELLS),
        "truncated": truncated,
    }


def recalc(source: Path, target: Path) -> dict:
    source = source.resolve()
    target = target.resolve()
    if source.suffix.lower() != ".xlsx" or target.suffix.lower() != ".xlsx":
        raise OfficeError("recalc currently requires .xlsx input and output")
    if source == target:
        raise OfficeError("input and output must be different paths")
    if not source.is_file():
        raise OfficeError(f"input is not a file: {source}")

    before = _scan_workbook(source, data_only=False)
    with tempfile.TemporaryDirectory(prefix="artifactflow-office-recalc-") as tmp:
        converted, result = _convert_to_temp(source, ".xlsx", Path(tmp))
        after_formulas = _scan_workbook(converted, data_only=False)
        cached = _scan_workbook(converted, data_only=True)
        if not before["truncated"] and not after_formulas["truncated"]:
            if after_formulas["formulas"] != before["formulas"]:
                raise OfficeError(
                    "formula count changed during LibreOffice save: "
                    f"{before['formulas']} -> {after_formulas['formulas']}"
                )
        _atomic_copy(converted, target)

    payload = {
        "ok": True,
        "operation": "recalc",
        "input": str(source),
        "output": str(target),
        "formulas": after_formulas["formulas"],
        "cached_errors": cached["cached_errors"],
        "cached_error_count": cached["cached_error_count"],
        "scan_truncated": before["truncated"] or after_formulas["truncated"] or cached["truncated"],
    }
    warning = result.stderr.strip()
    if warning:
        payload["warning"] = warning[-1000:]
    return payload


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--version", action="version", version=f"artifactflow-office {CLI_VERSION}")
    subparsers = parser.add_subparsers(dest="command", required=True)

    convert_parser = subparsers.add_parser("convert", help="convert an Office document to an explicit output path")
    convert_parser.add_argument("input")
    convert_parser.add_argument("output")

    render_parser = subparsers.add_parser("render", help="render Office/PDF pages to PNG files")
    render_parser.add_argument("input")
    render_parser.add_argument("outdir")
    render_parser.add_argument("--pages", help='1-based pages, e.g. "1-5,8"')

    recalc_parser = subparsers.add_parser("recalc", help="recalculate and save an .xlsx workbook")
    recalc_parser.add_argument("input")
    recalc_parser.add_argument("output")
    return parser


def main() -> None:
    args = _build_parser().parse_args()
    try:
        if args.command == "convert":
            payload = convert(Path(args.input), Path(args.output))
        elif args.command == "render":
            payload = render(Path(args.input), Path(args.outdir), args.pages)
        else:
            payload = recalc(Path(args.input), Path(args.output))
    except (OfficeError, OSError, ValueError) as exc:
        _emit({"ok": False, "operation": args.command, "error": str(exc)}, stream=sys.stderr)
        raise SystemExit(1) from None
    _emit(payload)


if __name__ == "__main__":
    main()
